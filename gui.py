import pandas as pd
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow,QMessageBox,QLineEdit, QFileDialog
from PyQt5 import QtCore
from ui import Ui_Dialog as mainUI
from ui1 import Ui_Dialog as firstUI
from ui2 import Ui_Dialog as secondUI
import os
from openpyxl import load_workbook
from time import sleep
from win32com.client import Dispatch
from PyQt5.QtCore import *
import threading
import pythoncom

pd.set_option('display.float_format',lambda x : '%.6f' % x)
pd.set_option('display.max_columns', None)

class Runthread(QtCore.QThread):
    # 通过类成员对象定义信号对象
    _signal = pyqtSignal(str,dict)

    def __init__(self,data):
        super(Runthread, self).__init__()
        self.path=data['path']
        self.input_num=data['input_num']

    def __del__(self):
        self.wait()

    def run(self):
        self.normal_list=[0.999904, 4.2e-05, 4.2e-05, 0.99991, 4.5e-05, 4.5e-05, 0.999916, 4.8e-05, 4.8e-05]
        self.normal_list1=[0.999904, '——', 8.4e-05, 0.99991, '——', 9e-05, 0.999916, '——', 9.6e-05]
        self.num_list=[4,5,6,8,11,12,13,14,15,17,18,19,20,22,23,24,25,26,27]
        res_list=[]
        path=self.path
        data=pd.DataFrame(pd.read_excel(path,sheet_name='总程序',usecols ='R:AW',nrows = 83))[72:]
        self.data=data

        #32列名
        all_list=data.columns.tolist()
        self.all_list=all_list
        for i in range(len(all_list)):
            keys=all_list[i]
            key=data[keys].tolist()
            del key[3]
            del key[6]
            if i in self.num_list:
                if key==self.normal_list1:
                    continue
                else:
                    res_list.append(i+1)
                    continue
            if key==self.normal_list:
                continue
            else:
                res_list.append(i+1)
                continue
        self.find_out1(res_list)

    def find_out1(self,res_list):
        all_result={}
        #决策组合 Sheet1
        path='决策组合.xlsx'
        data=pd.DataFrame(pd.read_excel(path,sheet_name='Sheet1',usecols ='A:D',nrows = 775))
        #措施 List
        cuoshi=data['决策措施'].tolist()
        #编号 List
        index=data['决策组合编号'].tolist()
        #花费 List
        cost=data['决策费用（万元）'].tolist()
        #事件 list
        cases=data['对应底事件'].tolist()
        #range(len(cases))
        dishijian=[[1],[2],[3],[5],[6],[7],[9],[11],[13,14,15,16],[30,31,32]]
        copy_di=[]
        for dishi in dishijian:
            for di in dishi:
                if di in res_list:
                    flag4=1
                    break
                copy_di.append(dishi)
                break
        cop_di=[]
        for keys in copy_di:
            for key in keys:
                cop_di.append(key)
        ca_list=cases[:]
        for op in range(len(cases)):
            case=cases[op].split('x')
            copy_case=cases[op]
            del case[0]
            for cas in case:
                if int(cas) in cop_di:
                    ca_list.remove(copy_case)
                    break

        for op in range(len(ca_list)):
            case=ca_list[op].split('x')
            fla_index=ca_list[op]
            del case[0]
            flag2=0
            #判断 不安全列表 是否在 决策组合中
            for cas in case:
                if int(cas) in res_list:
                    flag2=1
                    break
            if flag2:
                return_datas=[]
                for i in range(len(self.all_list)):
                    return_data=[]
                    if i+1 in res_list and str(i+1) in case:
                        if i in self.num_list:
                            return_data=self.normal_list1
                            return_datas.append(return_data)
                            continue
                        else:
                            return_data=self.normal_list
                            return_datas.append(return_data)
                            continue
                    else:
                        keys=self.all_list[i]
                        key=self.data[keys].tolist()
                        del key[3]
                        del key[6]
                        return_data=key
                        return_datas.append(return_data)
                        continue
                lp =self.change_value(return_datas)
                if lp:
                    pl=''
                    pl=cases.index(fla_index)
                    progress=str(op)+'/'+str(len(ca_list))
                    all_result[pl]=[cost[pl],cuoshi[pl]]
                else:
                    progress=str(op)+'/'+str(len(ca_list))
                self._signal.emit(progress,all_result)
            else:
                continue

    def change_value(self,return_datas):
        path='data.xlsx'
        x_list=['R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW']
        y_list=[74,75,76,78,79,80,82,83,84]
        wb=load_workbook(path)
        name_list = wb.get_sheet_names()
        # 根据 sheet 名字获得 sheet
        j=0
        for name in name_list:
            my_sheet = wb.get_sheet_by_name(name)
            if my_sheet.title=='总程序':
                break
            j+=1
        ws=wb.worksheets[j]
        #依次修改值
        for i in range(len(x_list)):
            for k in range(len(y_list)):
                name=x_list[i]+str(y_list[k])
                ws[name]=return_datas[i][k]
                k+=1
            i+=1
        wb.save(path)
        wb.close()
        self.just_open(path)
        wb1=load_workbook(path,data_only=True)
        ws1=wb1.worksheets[j]
        t1=round(ws1['A2'].value,6)
        t2=round(ws1['A3'].value,6)
        res=(t1+t2)*int(self.input_num)
        if res<=20:
            return True
        else:
            return False

    def just_open(self,filename):
        pythoncom.CoInitialize()
        path=os.getcwd()
        xlApp = Dispatch("Excel.Application")
        xlApp.Visible = False
        xlBook = xlApp.Workbooks.Open(path+'\\'+filename)
        xlBook.Save()
        xlBook.Close()

class First(QMainWindow,firstUI):
    flag=0
    flag1=0
    def __init__(self, parent=None):
        super(First, self).__init__(parent)
        self.setupUi(self)
        #选择文件
        self.file_button.clicked.connect(self.choice_file)
        #确认文件
        self.file_ensure_buttom.clicked.connect(self.ensure_choice_file)
        #取消文件
        self.file_cancel_buttom.clicked.connect(self.cancel_choice_file)
        #开始
        self.begin_button.clicked.connect(self.display)
        #是
        self.ensure_button.clicked.connect(self.show_result)
        #导出结果
        self.result_button.clicked.connect(self.export)

    def export(self):
        if not self.flag:
            return
        name=self.path.split('/')[-1].replace('xlsx','txt')
        with open(name,'w+') as f:
            _,risk,result=self.showdata()
            f.write('                          泵站配水间系统风险评价')
            f.write('\n')
            f.write('一.风险等级'+'\n')
            f.write('根据目前各个系统所处现状，整理分析各个评价细则可以得到如下结果:'+'\n')
            f.write('1.设备因素风险等级：                               '+risk[1]+'\n')
            f.write('2.管理人为因素风险等级:                           '+risk[2]+'\n')
            f.write('3.环境因素风险等级:                                  '+risk[3]+'\n')
            f.write('据此，可以得出整体系统的风险等级为:       '+result+'\n')
            f.write('二.风险管控措施'+'\n')
            f.write(self.text)
        QMessageBox.question(self,"导出结果","结果导出完成",QMessageBox.Yes|QMessageBox.No,QMessageBox.Yes)

    def choice_file(self):
        fileName, _ = QFileDialog.getOpenFileName(self,
                                    "选取文件",
                                    "./",
                                    "xlsx Files (*.xlsx)")   #设置文件扩展名过滤,注意用双分号间隔
        self.path=fileName
        filename=fileName.split('/')[-1]
        self.filename.setText(filename)

    def ensure_choice_file(self):
        if self.flag1:
            QMessageBox.critical(self,"错误信息","请点击取消按钮再次选择文件",QMessageBox.Yes|QMessageBox.No,QMessageBox.Yes)
            return
        try:
            print(self.path)
            self.flag1=1
        except:
            QMessageBox.critical(self,"错误信息","请选择文件",QMessageBox.Yes|QMessageBox.No,QMessageBox.Yes)


    def cancel_choice_file(self):
        self.flag1=0
        self.flag=0
        self.filename.setText('')
        self.result_text.setText('')
        self.result1_text.setText('')
        self.result2_text.setText('')
        self.result3_text.setText('')
        self.change_result_text.setText('')
        self.factor.setText('')


    def display(self):
        if not self.flag1:
             QMessageBox.critical(self,"错误信息","请选择文件或确认选择",QMessageBox.Yes|QMessageBox.No,QMessageBox.Yes)
             return
        self.flag=1
        factor,risk,result=self.showdata()
        self.result_text.setText(result)
        self.result1_text.setText(risk[1])
        self.result2_text.setText(risk[2])
        self.result3_text.setText(risk[3])

    def show_result(self):
        if self.flag==0:
            return
        data=self.solution()
        data=list(set(data))
        self.change_result_text.setText('低风险')
        i=0
        text=''
        for i in range(len(data)):
            text+=str(i+1)+':'+data[i]+'\n'
        self.text=text
        self.factor.setText(text)

    def solution(self):
        path=self.path
        data=pd.DataFrame(pd.read_excel(path,sheet_name='设备因素评价',usecols ='I,Q,V',nrows = 70))[3:]
        effect=data['Unnamed: 8'].tolist()
        step=data['Unnamed: 16'].tolist()
        choice=data['Unnamed: 21'].tolist()
        res=[]
        #根据输入获取解决措施
        i=0
        for key in choice:
            if key==1.0:
                if effect[i]=='危险':
                    res.append(step[i])
            i+=1
        return res

    def showdata(self):
        #FMEA Sheet3
        path=self.path
        data=pd.DataFrame(pd.read_excel(path,sheet_name='Sheet3',usecols ='C,D,L',nrows = 6))
        factor=data['Unnamed: 2'].tolist()
        risk=data['Unnamed: 3'].tolist()
        result=data['Unnamed: 11'].tolist()[1]
        return factor,risk,result


class Second(QMainWindow,secondUI):
    flag5=0
    flag1=0
    def __init__(self, parent=None):
        super(Second, self).__init__(parent)
        self.setupUi(self)
        self.ensure_button_3.clicked.connect(self.find_out)
        #选择文件
        self.file_button.clicked.connect(self.choice_file)
        #确认文件
        self.file_ensure_buttom.clicked.connect(self.ensure_choice_file)
        #取消文件
        self.file_cancel_buttom.clicked.connect(self.cancel_choice_file)
        #导出结果
        self.result_button.clicked.connect(self.export)

    def export(self):
        name=self.path.split('/')[-1].replace('xlsx','txt')
        with open(name,'w+') as f:
            f.write('                        注水管道风险评价')
            f.write('\n')
            f.write('一.风险等级'+'\n')
            f.write('系统的风险等级为:       '+self.text+'\n')
            f.write('二.风险管控措施'+'\n')
            f.write(self.cuoshiresult)
        QMessageBox.question(self,"导出结果","结果导出完成",QMessageBox.Yes|QMessageBox.No,QMessageBox.Yes)

    def choice_file(self):
        fileName, _ = QFileDialog.getOpenFileName(self,
                                    "选取文件",
                                    "./",
                                    "xlsx Files (*.xlsx)")   #设置文件扩展名过滤,注意用双分号间隔
        self.path=fileName
        filename=fileName.split('/')[-1]
        self.filename.setText(filename)

    def ensure_choice_file(self):
        try:
            print(self.path)
        except:
            QMessageBox.critical(self,"错误信息","请选择文件",QMessageBox.Yes|QMessageBox.No,QMessageBox.Yes)
            return
        self.input_num()


    def cancel_choice_file(self):
        self.filename.setText('')
        self.path=''

    def input_num(self):
        path=self.path
        data=pd.DataFrame(pd.read_excel(path,sheet_name='总程序',usecols ='F',nrows = 43))[41:42]
        self.input_num=data['Unnamed: 5'].tolist()[0]
        self.flag1=1
        self.evaluation()
        return

    def cancel(self):
        if self.flag5==1:
            QMessageBox.warning(self,"警告","请等待优化程序运行完成",QMessageBox.Yes|QMessageBox.No,QMessageBox.Yes)
            return
        self.input.setText('')
        self.flag1=0
        self.result.setText('')
        self.result_2.setText('')
        self.cuoshi.setText('')
        self.progressBar.setProperty("value",0)
        return

    def evaluation(self):
        #风险评估 总程序
        path=self.path
        if not os.path.exists(path):
            QMessageBox.critical(self,"错误信息","请检查目录下是否存在所需文件",QMessageBox.Yes|QMessageBox.No,QMessageBox.Yes)
        data=pd.DataFrame(pd.read_excel(path,sheet_name='总程序',usecols ='A',nrows = 2))
        data=data.values.tolist()
        t1=round(data[0][0],6)
        t2=round(data[1][0],6)
        res=(t1+t2)*int(self.input_num)
        text=''
        if res>=0 and res<=4:
            text='低'
        elif res>4 and res<=10:
            text='一般'
        elif res>10 and res<=20:
            text='较大'
        elif res>20 and res<=100:
            text='重大'
        self.text=text
        self.result.setText(text)

    def find_out(self):
        if not self.flag1:
            return
        if self.text=='低':
            QMessageBox.warning(self,"提示信息","评估风险为低，可以不改善",QMessageBox.Yes|QMessageBox.No,QMessageBox.Yes)
            return
        data={'input_num':self.input_num,'path':self.path}
        self.thread = Runthread(data)
        # 连接信号
        self.thread._signal.connect(self.call_backlog)  # 进程连接回传到GUI的事件
        # 开始线程
        self.flag5=1
        self.thread.start()

    def call_backlog(self, progress, all_result):
        progress=progress.split('/')
        if str(int(progress[0])+1)!=progress[1]:
            self.progressBar.setProperty("value", int(int(progress[0])*100/int(progress[1])))
        else:
            self.progressBar.setProperty("value", 100)
        if str(int(progress[0])+1)==progress[1]:
            self.flag5=0
            self.result_2.setText('低风险')
            min_value=99999
            cuoshi=''
            for _,value in all_result.items():
                if value[0]<min_value:
                    min_value=value[0]
                    cuoshi=value[1]
            self.cuoshi.setText(cuoshi)
            self.cuoshiresult=cuoshi

class Main(QMainWindow,mainUI):
    switch_window1 = QtCore.pyqtSignal() # 跳转信号
    switch_window2 = QtCore.pyqtSignal() # 跳转信号
    def __init__(self, parent=None):
        super(Main, self).__init__(parent)
        self.setupUi(self)
        self.left.clicked.connect(self.show_left)
        self.right.clicked.connect(self.show_right)

    def show_left(self):
        self.switch_window1.emit()
    def show_right(self):
        self.switch_window2.emit()

class Controller:
    def __init__(self):
        pass
    def show_main(self):
        self.main = Main()
        self.main.switch_window1.connect(self.show_left)
        self.main.switch_window2.connect(self.show_right)
        self.main.show()
    def show_left(self):
        self.left = First()
        self.main.close()
        self.left.show()
    def show_right(self):
        self.right = Second()
        self.main.close()
        self.right.show()

if __name__ == "__main__":
    #固定的，PyQt5程序都需要QApplication对象。sys.argv是命令行参数列表，确保程序可以双击运行
    app = QApplication(sys.argv)

    controller = Controller() # 控制器实例  
    controller.show_main() # 默认展示的是 hello 页面
    sys.exit(app.exec_())
